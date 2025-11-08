"""
Example Script: Processing Your Uploaded Files
Demonstrates how to use the TA Ops Audit system with your actual data

This script shows:
1. How to read the SharePoint audit tracker image
2. How to process the example requisition PDF
3. How to set up the pay transparency mapping
4. Complete end-to-end workflow
"""

import sys
from pathlib import Path
from ta_ops_audit_agent import (
    TAOpsAuditOrchestrator,
    GroundTruth,
    ExtractedData,
    ValidationStatus
)
from enhanced_corrector import EnhancedDocumentCorrector, SmartCorrectionAgent
from config_and_setup import TAOpsConfig, PayTransparencyLoader


def extract_ground_truth_from_sharepoint_image():
    """
    In production, this would read from SharePoint API
    For now, we'll manually create the ground truth based on the image
    
    From TA_Ops_Audit_Report.png, we can see:
    Job 614552: 29R-SoCal, $16.60-$26.75, Template=No, needs checking
    Job 614555: 27R-Seattle, $16.91-$24.15, All validations pass
    Job 614557: 27R-Seattle, $16.91-$24.15, All validations pass
    etc.
    """
    
    # Example ground truths extracted from the image
    ground_truths = [
        GroundTruth(
            job_requisition_number="614552",
            business_unit_name="29R-SoCal",
            min_pay_rate=16.60,
            max_pay_rate=26.75,
            facility="3044",
            state="CA"  # SoCal = California
        ),
        GroundTruth(
            job_requisition_number="614555",
            business_unit_name="27R-Seattle",
            min_pay_rate=16.91,
            max_pay_rate=24.15,
            facility="3424",
            state="WA"  # Seattle = Washington
        ),
        GroundTruth(
            job_requisition_number="614557",
            business_unit_name="27R-Seattle",
            min_pay_rate=16.91,
            max_pay_rate=24.15,
            facility="1864",
            state="WA"
        ),
        # Add more as needed...
    ]
    
    return ground_truths


def create_pay_transparency_mapping():
    """
    Create a sample pay transparency mapping
    In production, this would be loaded from your Excel file
    """
    
    transparency_map = {
        'CA': """
Pay Transparency (California):

Starting rates will be no less than the local minimum wage and may vary based on 
things like location, experience, qualifications, and the terms of any applicable 
collective bargaining agreement. California law requires that we provide a pay scale 
as follows: The budgeted range listed in this posting reflects what we reasonably 
expect to pay for this position. Dependent on length of service, hours worked and any 
applicable collective bargaining agreement, benefits may include medical, dental, 
vision, disability and life insurance, sick pay, PTO/Vacation pay and retirement 
benefits (pension and/or 401(k) eligibility). This is an entry level position with 
advancement opportunity. Applications are accepted on an on-going basis.
        """,
        
        'WA': """
Pay Transparency (Washington):

Starting rates will be no less than the local minimum wage and may vary based on 
things like location, experience, qualifications, and the terms of any applicable 
collective bargaining agreement. Washington law requires disclosure of salary or 
wage range: The budgeted range listed in this posting reflects what we reasonably 
expect to pay for this position. Dependent on length of service, hours worked and 
any applicable collective bargaining agreement, benefits may include medical, dental, 
vision, disability and life insurance, sick pay, PTO/Vacation pay and retirement 
benefits (pension and/or 401(k) eligibility). This is an entry level position with 
advancement opportunity. Applications are accepted on an on-going basis.
        """,
        
        'IL': """
Pay Transparency (Illinois):

Starting rates will be no less than the local minimum wage and may vary based on 
things like location, experience, qualifications, and the terms of any applicable 
collective bargaining agreement. Dependent on length of service, hours worked and 
any applicable collective bargaining agreement, benefits may include medical, dental, 
vision, disability and life insurance, sick pay, PTO/Vacation pay and retirement 
benefits (pension and/or 401(k) eligibility). This is an entry level position with 
advancement opportunity. Applications are accepted on an on-going basis.
        """,
        
        # Add more states as needed
        'CO': """Pay Transparency (Colorado): ...""",
        'NY': """Pay Transparency (New York): ...""",
    }
    
    return transparency_map


def process_example_requisition():
    """
    Process the example requisition that was uploaded
    Job 651640 - Retail Sales and Store Support - Downers Grove, IL
    """
    
    print("\n" + "="*70)
    print("EXAMPLE: Processing Job Requisition 651640")
    print("="*70 + "\n")
    
    # Step 1: Define ground truth for this job
    # Based on the PDF, we see Min=$14.75, Max=$15.00
    ground_truth = GroundTruth(
        job_requisition_number="651640",
        business_unit_name="Unknown",  # Not in the SharePoint image
        min_pay_rate=14.75,  # From the PDF
        max_pay_rate=15.00,   # From the PDF
        facility="1148",      # From the address
        state="IL"            # From the address
    )
    
    print("[1] Ground Truth Defined:")
    print(f"    Job ID: {ground_truth.job_requisition_number}")
    print(f"    Min Pay: ${ground_truth.min_pay_rate:.2f}")
    print(f"    Max Pay: ${ground_truth.max_pay_rate:.2f}")
    print(f"    State: {ground_truth.state}")
    print()
    
    # Step 2: In a real scenario, you would initialize the orchestrator
    # For this example, we'll simulate the process
    
    print("[2] Simulating Document Extraction...")
    print("    ✓ Extracted Job ID: 651640")
    print("    ✓ Extracted Banner: Jewel Osco")
    print("    ✓ Extracted Location: 1148 OGDEN AVE, DOWNERS GROVE, IL, 60515")
    print("    ✓ Extracted Min Rate: $14.75")
    print("    ✓ Extracted Max Rate: $15.00")
    print("    ✓ Found role-specific rates:")
    print("        - Meat Associate: $13.50 - $16.55")
    print("        - Floral Associate: $13.75 - $14.00")
    print("        - Grocery Associate: $13.75 - $14.00")
    print()
    
    # Step 3: Validation
    print("[3] Running Validations...")
    
    validations = {
        'Template Structure': 'PASS ✓',
        'Min Pay Rate': 'PASS ✓',
        'Max Pay Rate': 'PASS ✓',
        'Job Description': 'PASS ✓',
        'Dollar Signs': 'PASS ✓',
        'Pay Transparency (IL)': 'CHECKING...'
    }
    
    for check, status in validations.items():
        print(f"    {check}: {status}")
    
    print()
    
    # Step 4: Check role-specific rates
    print("[4] Validating Role-Specific Pay Rates...")
    print("    Checking if rates fall within $14.75 - $15.00 range:")
    print("    ⚠ Issue: Meat Associate max ($16.55) exceeds max rate ($15.00)")
    print("    ✓ Floral Associate range within limits")
    print("    ✓ Grocery Associate range within limits")
    print()
    
    # Step 5: Correction plan
    print("[5] Generating Correction Plan...")
    print("    Priority: MEDIUM")
    print("    Correction 1: Adjust Meat Associate max rate from $16.55 to $15.00")
    print("    Correction 2: Verify Pay Transparency text for Illinois")
    print()
    
    # Step 6: Apply corrections
    print("[6] Applying Corrections...")
    print("    ✓ Corrected Meat Associate rate: $13.50 - $15.00")
    print("    ✓ Verified Pay Transparency text")
    print("    ✓ Generated corrected document: 651640_CORRECTED.docx")
    print()
    
    # Step 7: Report generation
    print("[7] Generating Audit Report...")
    print("    ✓ Report saved: 651640_AUDIT_REPORT.txt")
    print("    ✓ SharePoint would be updated with validation results")
    print()
    
    print("="*70)
    print("PROCESSING COMPLETE!")
    print("="*70)
    print()
    print("Summary:")
    print("  - Issues Found: 1")
    print("  - Corrections Applied: 1")
    print("  - Status: APPROVED with corrections")
    print()


def demonstrate_batch_processing():
    """
    Demonstrate how batch processing would work with multiple requisitions
    """
    
    print("\n" + "="*70)
    print("EXAMPLE: Batch Processing Multiple Requisitions")
    print("="*70 + "\n")
    
    # Get ground truths from SharePoint image
    ground_truths = extract_ground_truth_from_sharepoint_image()
    
    print(f"[1] Loaded {len(ground_truths)} requisitions from SharePoint tracker")
    print()
    
    # Simulate processing each one
    results_summary = []
    
    for i, gt in enumerate(ground_truths, 1):
        print(f"[{i}/{len(ground_truths)}] Processing Job {gt.job_requisition_number}...")
        
        # Simulate processing
        issues = 0
        corrections = 0
        
        # Based on the image, job 614552 has issues
        if gt.job_requisition_number == "614552":
            issues = 2
            corrections = 2
            status = "CORRECTED"
        else:
            status = "PASSED"
        
        results_summary.append({
            'job_id': gt.job_requisition_number,
            'issues': issues,
            'corrections': corrections,
            'status': status
        })
        
        print(f"    Status: {status}")
        if issues > 0:
            print(f"    Issues: {issues}, Corrections: {corrections}")
        print()
    
    # Print summary
    print("="*70)
    print("BATCH PROCESSING COMPLETE!")
    print("="*70)
    print()
    print("Summary:")
    print(f"  Total Processed: {len(results_summary)}")
    print(f"  Passed: {sum(1 for r in results_summary if r['status'] == 'PASSED')}")
    print(f"  Corrected: {sum(1 for r in results_summary if r['status'] == 'CORRECTED')}")
    print(f"  Total Corrections: {sum(r['corrections'] for r in results_summary)}")
    print()


def demonstrate_real_world_workflow():
    """
    Show the complete real-world workflow
    """
    
    print("\n" + "="*70)
    print("REAL-WORLD WORKFLOW DEMONSTRATION")
    print("="*70 + "\n")
    
    print("Step 1: System Setup")
    print("─────────────────────")
    print("  ✓ Load configuration from config.yaml")
    print("  ✓ Connect to SharePoint API")
    print("  ✓ Load pay transparency mappings")
    print("  ✓ Initialize Claude API client")
    print()
    
    print("Step 2: Data Collection")
    print("───────────────────────")
    print("  ✓ Fetch ground truth from SharePoint 'TA Ops Audit Report' list")
    print("  ✓ Scan PDF directory for requisition documents")
    print("  ✓ Match PDFs to job requisition numbers")
    print()
    
    print("Step 3: Processing Pipeline")
    print("────────────────────────────")
    print("  For each requisition:")
    print("    → Agent 1: Extract data from PDF using Claude")
    print("    → Agent 2: Validate against ground truth")
    print("    → Agent 3: Apply corrections if needed")
    print("    → Agent 4: Update SharePoint & generate report")
    print()
    
    print("Step 4: Quality Control")
    print("───────────────────────")
    print("  ✓ Human review of flagged requisitions")
    print("  ✓ Approval of corrected documents")
    print("  ✓ Final publish to Oracle HCM")
    print()
    
    print("Step 5: Reporting & Monitoring")
    print("───────────────────────────────")
    print("  ✓ Daily summary emails")
    print("  ✓ Slack notifications for errors")
    print("  ✓ Dashboard metrics update")
    print("  ✓ Audit trail archive")
    print()


def create_sample_pay_transparency_excel():
    """
    Create a sample Excel file for pay transparency mapping
    """
    
    import openpyxl
    
    print("\n[Creating Sample Pay Transparency Excel File]")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pay Transparency"
    
    # Headers
    ws['A1'] = "State Code"
    ws['B1'] = "Pay Transparency Text"
    
    # Make headers bold
    for cell in ws['1:1']:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Sample data
    data = [
        ('CA', 'Starting rates will be no less than the local minimum wage and may vary based on things like location, experience, qualifications, and the terms of any applicable collective bargaining agreement. California law requires that we provide a pay scale as follows: The budgeted range listed in this posting reflects what we reasonably expect to pay for this position.'),
        ('WA', 'Starting rates will be no less than the local minimum wage and may vary based on things like location, experience, qualifications, and the terms of any applicable collective bargaining agreement. Washington law requires disclosure of salary or wage range: The budgeted range listed in this posting reflects what we reasonably expect to pay for this position.'),
        ('IL', 'Starting rates will be no less than the local minimum wage and may vary based on things like location, experience, qualifications, and the terms of any applicable collective bargaining agreement.'),
        ('CO', 'Colorado law requires disclosure of salary or wage range: The budgeted range listed in this posting reflects what we reasonably expect to pay for this position.'),
        ('NY', 'New York law requires disclosure of salary or wage range: The budgeted range listed in this posting reflects what we reasonably expect to pay for this position.'),
    ]
    
    for i, (state, text) in enumerate(data, start=2):
        ws[f'A{i}'] = state
        ws[f'B{i}'] = text
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 100
    
    # Save file
    output_path = '/home/claude/pay_transparency_sample.xlsx'
    wb.save(output_path)
    
    print(f"  ✓ Created: {output_path}")
    print(f"  ✓ Contains {len(data)} state mappings")
    print()
    
    return output_path


def main():
    """
    Main function to run all demonstrations
    """
    
    print("\n" + "="*70)
    print(" TA OPS AUDIT SYSTEM - DEMONSTRATION")
    print(" Example Using Your Uploaded Files")
    print("="*70)
    
    # Create sample files
    print("\n[Setup Phase]")
    pay_transparency_file = create_sample_pay_transparency_excel()
    
    # Run demonstrations
    print("\n" + "="*70)
    print(" DEMONSTRATION 1: Single Requisition Processing")
    print("="*70)
    process_example_requisition()
    
    input("\nPress Enter to continue to Batch Processing demo...")
    
    demonstrate_batch_processing()
    
    input("\nPress Enter to see Real-World Workflow...")
    
    demonstrate_real_world_workflow()
    
    # Final instructions
    print("\n" + "="*70)
    print(" NEXT STEPS FOR PRODUCTION DEPLOYMENT")
    print("="*70 + "\n")
    
    print("1. Complete Configuration:")
    print("   → Add your Anthropic API key to config.yaml")
    print("   → Set up SharePoint credentials")
    print("   → Update file paths to your environment")
    print()
    
    print("2. Prepare Data:")
    print("   → Export SharePoint 'TA Ops Audit Report' list")
    print("   → Collect all requisition PDFs in one directory")
    print("   → Create complete pay transparency Excel file")
    print()
    
    print("3. Test Run:")
    print("   → python config_and_setup.py setup")
    print("   → python config_and_setup.py single --pdf test.pdf --job-id TEST001")
    print("   → Review generated outputs")
    print()
    
    print("4. Production Deployment:")
    print("   → python config_and_setup.py batch --pdf-dir /path/to/pdfs")
    print("   → Schedule with cron/Task Scheduler for automation")
    print("   → Set up monitoring and alerts")
    print()
    
    print("="*70)
    print(" For questions or support:")
    print(" - Review README.md for complete documentation")
    print(" - Contact: GenAI Team at Philtech Inc.")
    print("="*70 + "\n")


if __name__ == "__main__":
    main()

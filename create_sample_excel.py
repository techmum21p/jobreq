"""
Generate Sample Pay Transparency Excel File
"""

import openpyxl

def create_pay_transparency_excel():
    """Create a sample Excel file for pay transparency mapping"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pay Transparency"
    
    # Headers
    ws['A1'] = "State Code"
    ws['B1'] = "Pay Transparency Text"
    
    # Make headers bold
    for cell in ws['1:1']:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Sample data for different states
    data = [
        ('CA', 'Starting rates will be no less than the local minimum wage and may vary based on things like location, experience, qualifications, and the terms of any applicable collective bargaining agreement. California law requires that we provide a pay scale as follows: The budgeted range listed in this posting reflects what we reasonably expect to pay for this position. Dependent on length of service, hours worked and any applicable collective bargaining agreement, benefits may include medical, dental, vision, disability and life insurance, sick pay, PTO/Vacation pay and retirement benefits (pension and/or 401(k) eligibility). This is an entry level position with advancement opportunity. Applications are accepted on an on-going basis.'),
        
        ('WA', 'Starting rates will be no less than the local minimum wage and may vary based on things like location, experience, qualifications, and the terms of any applicable collective bargaining agreement. Washington law requires disclosure of salary or wage range: The budgeted range listed in this posting reflects what we reasonably expect to pay for this position. Dependent on length of service, hours worked and any applicable collective bargaining agreement, benefits may include medical, dental, vision, disability and life insurance, sick pay, PTO/Vacation pay and retirement benefits (pension and/or 401(k) eligibility). This is an entry level position with advancement opportunity. Applications are accepted on an on-going basis.'),
        
        ('IL', 'Starting rates will be no less than the local minimum wage and may vary based on things like location, experience, qualifications, and the terms of any applicable collective bargaining agreement. Dependent on length of service, hours worked and any applicable collective bargaining agreement, benefits may include medical, dental, vision, disability and life insurance, sick pay, PTO/Vacation pay and retirement benefits (pension and/or 401(k) eligibility). This is an entry level position with advancement opportunity. Applications are accepted on an on-going basis.'),
        
        ('CO', 'Starting rates will be no less than the local minimum wage and may vary based on things like location, experience, qualifications, and the terms of any applicable collective bargaining agreement. Colorado law requires disclosure of salary or wage range: The budgeted range listed in this posting reflects what we reasonably expect to pay for this position. Dependent on length of service, hours worked and any applicable collective bargaining agreement, benefits may include medical, dental, vision, disability and life insurance, sick pay, PTO/Vacation pay and retirement benefits (pension and/or 401(k) eligibility). This is an entry level position with advancement opportunity. Applications are accepted on an on-going basis.'),
        
        ('NY', 'Starting rates will be no less than the local minimum wage and may vary based on things like location, experience, qualifications, and the terms of any applicable collective bargaining agreement. New York law requires disclosure of salary or wage range: The budgeted range listed in this posting reflects what we reasonably expect to pay for this position. Dependent on length of service, hours worked and any applicable collective bargaining agreement, benefits may include medical, dental, vision, disability and life insurance, sick pay, PTO/Vacation pay and retirement benefits (pension and/or 401(k) eligibility). This is an entry level position with advancement opportunity. Applications are accepted on an on-going basis.'),
        
        ('OR', 'Starting rates will be no less than the local minimum wage and may vary based on things like location, experience, qualifications, and the terms of any applicable collective bargaining agreement. Dependent on length of service, hours worked and any applicable collective bargaining agreement, benefits may include medical, dental, vision, disability and life insurance, sick pay, PTO/Vacation pay and retirement benefits (pension and/or 401(k) eligibility). This is an entry level position with advancement opportunity. Applications are accepted on an on-going basis.'),
        
        ('TX', 'Starting rates will be no less than the local minimum wage and may vary based on things like location, experience, qualifications, and the terms of any applicable collective bargaining agreement. Dependent on length of service, hours worked and any applicable collective bargaining agreement, benefits may include medical, dental, vision, disability and life insurance, sick pay, PTO/Vacation pay and retirement benefits (pension and/or 401(k) eligibility). This is an entry level position with advancement opportunity. Applications are accepted on an on-going basis.'),
        
        ('AZ', 'Starting rates will be no less than the local minimum wage and may vary based on things like location, experience, qualifications, and the terms of any applicable collective bargaining agreement. Dependent on length of service, hours worked and any applicable collective bargaining agreement, benefits may include medical, dental, vision, disability and life insurance, sick pay, PTO/Vacation pay and retirement benefits (pension and/or 401(k) eligibility). This is an entry level position with advancement opportunity. Applications are accepted on an on-going basis.'),
    ]
    
    # Add data rows
    for i, (state, text) in enumerate(data, start=2):
        ws[f'A{i}'] = state
        ws[f'B{i}'] = text
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 100
    
    # Save file
    output_path = 'pay_transparency_by_state.xlsx'
    wb.save(output_path)
    
    print(f"✓ Created: {output_path}")
    print(f"✓ Contains {len(data)} state mappings")
    
    return output_path

if __name__ == "__main__":
    create_pay_transparency_excel()
